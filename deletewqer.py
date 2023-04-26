import openpyxl
# import json
# dataframe = {
#   "type": "service_account",
#   "project_id": "leafy-racer-383704",
#   "private_key_id": "c782dad2d83643226b2701f846a2ff5b68c574ce",
#   "private_key": "-----BEGIN PRIVATE KEY-----\nMIIEvAIBADANBgkqhkiG9w0BAQEFAASCBKYwggSiAgEAAoIBAQCp0+nT8UroYl/G\nRcKyABWi89kn/p74ax85qWWRQZCaED6Te7wbQF3S4iX1JRTEpcVCttCf7aSw4sQE\nOA+LBR8Gldcf9xhxLNM83vKJbxaaSUfozlKklOFapmaKgajjSlUqPNhrglyW7S/i\nSt54IeIXdX2JlHsteMxa8QP1DPkiJE1Wixq5WeDhL6sTzpeBK85Q+mPpCDEfSzi2\nGnZ5/ILI20t9+y7sJ+DbN0kabRLhBNmRZyAye8/dRnQjlMym+1GU+9PB74Kx/xtu\nG7v/0DNScionvAnOp/up8jVDZJOqWLvgOGK8G2jsaipO0gwRvLK6ykOgni20rhNM\nEU0KnpBbAgMBAAECggEABYFqnCwlxIc9gANoJQMefT9KGlZr0U8PuEigk7PvbyWS\neCkgipwGLqPJA6IYthAcGmYsBaZ1NCYCoSxifJxFSp5cigG/SucsHFYR8/8KjL5O\nOxYToct3ZTJUMq5/ZIojw3aTgYpWKOJplS/eKtpv81i9iB9GWwDTU/adegbaPY9B\nzLwkS160DvRp6H5EdijYr+FkCmhg9FVClq4N+cUoz30y7Nkxg5oElRfMg1olJk9v\nW415EaeojSE5dUA+dkVcPg+tb6zwpR7hLNOTHmO46aBQuFUNp7Xs+JZSZFmiWohJ\nGqxgn1lye6HvFIrmSps09FfihHUEdiAR/9c06jUvgQKBgQDSWR/Mqo4IkyA026gr\neoah9A5O9S8xd7MNiacQRxYgTYXEsNDOFSV32BvwQSiFkjEzhFwo8f8hBy2GhC8k\nplFINsGwUkOeiM8Vn3N2QA6Gf/vWyp5cxxxfdRMle9Vi4j2H0FUNLrep6nkukoMw\nuvaIrvgoKS1ly1xYToEw9EtvmwKBgQDOr328ohA7xvRkNGrwnQ5Vjm+AVhh0eLJZ\nZu7ZlIgKcYyEmk2XR4KXxjY3Ya7pQv8EnZU3mp/wQ/17N52yNOIf5dDW8+dT/tTM\nOQjqJ4U5LcrxRseY1tH36ZNxTFDPjDE5okDfpXraU2q3VZoc3+IyGeO9fSK8NNQc\n/IB5nMJOQQKBgCStG4lRF7DRDmWJAzqJg54Lcs87mv3yFAtBPBrveiEpwzA3cEDu\nmPjbeouaoNkhU2jbJEKTZ1u3560zsRCf2z5PJdh7j4S97/WiNrygkKfM+vlJnPEq\nWn0eabc4b01tizL6FUpWIfoDgxumui8t+jx83ZILE/pSyggvsAD8J4cPAoGAbJxt\nv52OWEtRjAK7m5t8t9r0nwUfLr1qQ27C3xWdc5d91wTa9z890hH31ZyAdNWe+0bv\n7dtUZsbjwkdUmBwQI+D5oKtlCY+lNXhXRezkpkp5Fdcy9g8VIWjd8TDAWoB6MaYg\n3yWvbGFfd94E8V15GQjf02iZwzsjp+PMg5urw8ECgYBz4gTsrig07bO/rldcZbFT\nXTA2JId1pFarGFMxslHwi8bQcEcjrF3FTpD5yjdK0KtpT3haiaF1kdhpmytZz8B6\nrJ2PhLw96HkRbcNQujHEi9Q7LdErX8uGepQSu+i7dzeyefto2PvH/RDfwAap9ZIR\nQsJTrBfLlp2EmjfawAGpqg==\n-----END PRIVATE KEY-----\n",
#   "client_email": "credtype@leafy-racer-383704.iam.gserviceaccount.com",
#   "client_id": "100845028381799904064",
#   "auth_uri": "https://accounts.google.com/o/oauth2/auth",
#   "token_uri": "https://oauth2.googleapis.com/token",
#   "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
#   "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/credtype%40leafy-racer-383704.iam.gserviceaccount.com"
# }
# cont_str = json.dumps(dataframe)
# print(type(cont_str))

msglist1 = {'Serial Number': 'NULL', 'Battery Code': 'NULL', 'BATTERY_VOLTAGE': 'NULL', 'BATTERY_CURRENT': 'NULL', 'SOC': 'NULL', 'MAX_CELL_VOLTAGE': 'NULL', 'No_Of_Max_Cell_Voltage': 'NULL', 'MIN_CELL_VOLTAGE': 'NULL', 'No_Of_Min_Cell_Voltage': 'NULL', 'MAX_CELL_TEMPERATURE': 'NULL', 'Max_Temp_CellNo': 'NULL', 'MIN_CELL_TEMPERATURE': 'NULL', 'Min_Temp_CellNo': 'NULL', 'Charge_Discharge_Status': 'NULL', 'Charge_MOS_Tube_Status': 'NULL', 'Discharge_MOS_Tube_Status': 'NULL', 'BMS_Life': 'NULL', 'Residual_Capacity': 'NULL', 'NUMBER_OF_CELLS': 'NULL', 'NUMBER_OF_TEMPERATURE_SENSORS': 'NULL', 'Charger_Status': 'NULL', 'Load_Status': 'NULL', 'Charge_Discharge_Cycles': 'NULL', 'Frame_Number': 'NULL', 'CELL_1_VOLTAGE': 'NULL', 'CELL_4_VOLTAGE': 'NULL', 'CELL_7_VOLTAGE': 'NULL', 'CELL_10_VOLTAGE': 'NULL', 'CELL_13_VOLTAGE': 'NULL', 'CELL_16_VOLTAGE': 'NULL', 'CELL_19_VOLTAGE': 'NULL', 'CELL_2_VOLTAGE': 'NULL', 'CELL_5_VOLTAGE': 'NULL', 'CELL_8_VOLTAGE': 'NULL', 'CELL_11_VOLTAGE': 'NULL', 'CELL_14_VOLTAGE': 'NULL', 'CELL_17_VOLTAGE': 'NULL', 'CELL_20_VOLTAGE': 'NULL', 'CELL_3_VOLTAGE': 'NULL', 'CELL_6_VOLTAGE': 'NULL', 'CELL_9_VOLTAGE': 'NULL', 'CELL_12_VOLTAGE': 'NULL', 'CELL_15_VOLTAGE': 'NULL', 'CELL_18_VOLTAGE': 'NULL', 'Temp_Frame_No': 'NULL', 'TEMPERATURE_SENSOR_1': 'NULL', 'TEMPERATURE_SENSOR_2': 'NULL', 'TEMPERATURE_SENSOR_3': 'NULL', 'TEMPERATURE_SENSOR_4': 'NULL', 'Discharge_Switch_Status': 'NULL', 'Charging_Switch_Status': 'NULL', 'Response_capacity': 'NULL', 'Response_voltage': 'NULL', 'Read_capacity': 'NULL', 'Read_voltage': 'NULL', 'sum_volt_high_l1': 'NULL', 'sum_volt_high_l2': 'NULL', 'sum_volt_low_l1': 'NULL', 'sum_volt_low_l2': 'NULL', 'charge_current_l1': 'NULL', 'charge_current_l2': 'NULL', 'discharge_current_l1': 'NULL', 'discharge_current_l2': 'NULL', 'cell_volt_high_l1': 'NULL', 'cell_volt_high_l2': 'NULL', 'cell_volt_low_l1': 'NULL', 'cell_high_low_l2': 'NULL', 'volt_diff_large_l1': 'NULL', 'volt_diff_large_l2': 'NULL', 
'temp_diff_large_l1': 'NULL', 'temp_diff_large_l2': 'NULL', 'soc_high_l1': 'NULL', 'soc_high_l2': 'NULL', 'soc_low_l1': 'NULL', 'soc_low_l2': 'NULL', 'charge_temp_high_l1': 'NULL', 'charge_temp_high_l2': 'NULL', 'charge_temp_low_l1': 'NULL', 'charge_temp_low_l2': 'NULL', 'discharge_temp_high_l1': 'NULL', 'discharge_temp_high_l2': 'NULL', 'discharge_temp_low_l1': 'NULL', 'discharge_temp_low_l2': 'NULL', 'balance_start_volt': 'NULL', 'balance_start_diff_volt': 'NULL', 'battery_type': 'NULL', 'battery_operation_mode': 'NULL', 'year': 'NULL', 'month': 'NULL', 'date': 'NULL', 'time_sleep': 'NULL', 'current_wave': 'NULL', 'Frame_no_battery_code': 'NULL', 'String_1': 'NULL', 'String_8': 'NULL', 'String_15': 'NULL', 'String_22': 'NULL', 'String_29': 'NULL', 'String_2': 'NULL', 'String_9': 'NULL', 'String_16': 'NULL', 'String_23': 'NULL', 'String_30': 'NULL', 'String_3': 'NULL', 'String_10': 'NULL', 'String_17': 'NULL', 'String_24': 'NULL', 'String_31': 'NULL', 'String_4': 'NULL', 'String_11': 'NULL', 'String_18': 'NULL', 'String_25': 'NULL', 'String_32': 'NULL', 'String_5': 'NULL', 'String_12': 'NULL', 'String_19': 'NULL', 'String_26': 'NULL', 'String_33': 'NULL', 'String_6': 'NULL', 'String_13': 'NULL', 'String_20': 'NULL', 'String_27': 'NULL', 'String_34': 'NULL', 'String_7': 'NULL', 'String_14': 'NULL', 
'String_21': 'NULL', 'String_28': 'NULL', 'String_35': 'NULL', 'respons_bt_code': 'NULL'}
msglist2 = {'Sum_Volt_Low_Level_2': 'NULL', 'Sum_Volt_Low_Level_1': 'NULL', 'Sum_Volt_High_Level_2': 'NULL', 'Sum_Volt_High_Level_1': 'NULL', 'Cell_Volt_Low_Level_2': 'NULL', 'Cell_Volt_Low_Level_1': 'NULL', 'Cell_Volt_High_Level_2': 'NULL', 'Cell_Volt_High_Level_1': 'NULL', 'Dischg_Temp_Low_Level_2': 'NULL', 'Dischg_Temp_Low_Level_1': 'NULL', 'Dischg_Temp_High_Level_2': 'NULL', 'Dischg_Temp_High_Level_1': 'NULL', 'Chg_Temp_Low_Level_2': 'NULL', 'Chg_Temp_Low_Level_1': 'NULL', 'Chg_Temp_High_Level_2': 'NULL', 'Chg_Temp_High_Level_1': 'NULL', 'SOC_Low_Level_2': 'NULL', 'SOC_Low_Level_1': 'NULL', 'SOC_High_Level_2': 'NULL', 'SOC_High_Level_1': 'NULL', 'Dischg_Overcurrent_Level_2': 'NULL', 'Dischg_Overcurrent_Level_1': 'NULL', 'Chg_Overcurrent_Level_2': 'NULL', 'Chg_Overcurrent_Level_1': 'NULL', 'Diff_Temp_Level_2': 'NULL', 'Diff_Temp_Level_1': 'NULL', 'Diff_Volt_Level_2': 'NULL', 'Diff_Volt_Level_1': 'NULL'}
keylist = ['Serial Number', 'Battery Code', 'BATTERY_VOLTAGE', 'BATTERY_CURRENT', 'SOC', 'MAX_CELL_VOLTAGE', 'No_Of_Max_Cell_Voltage', 'MIN_CELL_VOLTAGE', 'No_Of_Min_Cell_Voltage', 'MAX_CELL_TEMPERATURE', 'Max_Temp_CellNo', 'MIN_CELL_TEMPERATURE', 'Min_Temp_CellNo', 'Charge_Discharge_Status', 'Charge_MOS_Tube_Status', 'Discharge_MOS_Tube_Status', 'BMS_Life', 'Residual_Capacity', 'NUMBER_OF_CELLS', 'NUMBER_OF_TEMPERATURE_SENSORS', 'Charger_Status', 'Load_Status', 'Charge_Discharge_Cycles', 'Frame_Number', 'CELL_1_VOLTAGE', 'CELL_4_VOLTAGE', 'CELL_7_VOLTAGE', 'CELL_10_VOLTAGE', 'CELL_13_VOLTAGE', 'CELL_16_VOLTAGE', 'CELL_19_VOLTAGE', 'CELL_2_VOLTAGE', 'CELL_5_VOLTAGE', 'CELL_8_VOLTAGE', 'CELL_11_VOLTAGE', 'CELL_14_VOLTAGE', 'CELL_17_VOLTAGE', 'CELL_20_VOLTAGE', 'CELL_3_VOLTAGE', 'CELL_6_VOLTAGE', 'CELL_9_VOLTAGE', 'CELL_12_VOLTAGE', 'CELL_15_VOLTAGE', 'CELL_18_VOLTAGE', 'Temp_Frame_No', 'TEMPERATURE_SENSOR_1', 'TEMPERATURE_SENSOR_2', 'TEMPERATURE_SENSOR_3', 'TEMPERATURE_SENSOR_4', 'Discharge_Switch_Status', 'Charging_Switch_Status', 'Response_capacity', 'Response_voltage', 'Read_capacity', 'Read_voltage', 'sum_volt_high_l1', 'sum_volt_high_l2', 'sum_volt_low_l1', 'sum_volt_low_l2', 'charge_current_l1', 'charge_current_l2', 'discharge_current_l1', 'discharge_current_l2', 'cell_volt_high_l1', 'cell_volt_high_l2', 'cell_volt_low_l1', 'cell_high_low_l2', 'volt_diff_large_l1', 'volt_diff_large_l2', 'temp_diff_large_l1', 'temp_diff_large_l2', 'soc_high_l1', 'soc_high_l2', 'soc_low_l1', 'soc_low_l2', 'charge_temp_high_l1', 'charge_temp_high_l2', 'charge_temp_low_l1', 'charge_temp_low_l2', 'discharge_temp_high_l1', 'discharge_temp_high_l2', 'discharge_temp_low_l1', 'discharge_temp_low_l2', 'balance_start_volt', 'balance_start_diff_volt', 'battery_type', 'battery_operation_mode', 'year', 'month', 'date', 'time_sleep', 'current_wave', 'Frame_no_battery_code', 'String_1', 'String_8', 'String_15', 'String_22', 'String_29', 'String_2', 'String_9', 'String_16', 'String_23', 'String_30', 'String_3', 'String_10', 'String_17', 'String_24', 'String_31', 'String_4', 'String_11', 'String_18', 'String_25', 'String_32', 'String_5', 'String_12', 'String_19', 'String_26', 'String_33', 'String_6', 'String_13', 'String_20', 'String_27', 'String_34', 'String_7', 'String_14', 'String_21', 'String_28', 'String_35', 'respons_bt_code', 'Sum_Volt_Low_Level_2', 'Sum_Volt_Low_Level_1', 'Sum_Volt_High_Level_2', 'Sum_Volt_High_Level_1', 'Cell_Volt_Low_Level_2', 'Cell_Volt_Low_Level_1', 'Cell_Volt_High_Level_2', 'Cell_Volt_High_Level_1', 'Dischg_Temp_Low_Level_2', 'Dischg_Temp_Low_Level_1', 'Dischg_Temp_High_Level_2', 'Dischg_Temp_High_Level_1', 'Chg_Temp_Low_Level_2', 'Chg_Temp_Low_Level_1', 'Chg_Temp_High_Level_2', 'Chg_Temp_High_Level_1', 'SOC_Low_Level_2', 'SOC_Low_Level_1', 'SOC_High_Level_2', 'SOC_High_Level_1', 'Dischg_Overcurrent_Level_2', 'Dischg_Overcurrent_Level_1', 'Chg_Overcurrent_Level_2', 'Chg_Overcurrent_Level_1', 'Diff_Temp_Level_2', 'Diff_Temp_Level_1', 'Diff_Volt_Level_2', 'Diff_Volt_Level_1']

print(len(msglist1),len(msglist2),len(keylist))

# for key, value in msglist1.items():
#     msglist1[key] = 45
# for key, value in msglist2.items():
#     msglist1[key] = 45

test = 'Null'
batterycode = ""
Serial_number = 'Test'
bms_battery_code = 'Test2'
dl_path = "Test_Data_Logging_.xlsx"
# dl_path = f"Data_Logging.zip\Data_Logging.xlsx"
dl_obj = openpyxl.load_workbook(dl_path)
dl_sheet_obj = dl_obj.active
dl_no_of_rows = dl_sheet_obj.max_row
print('These are the max number of rows appended', dl_no_of_rows)
# print(len(msglist1), len(keylist), len(msglist2))
for i in range(0, len(msglist1)):
    dl_sheet_obj.cell(row = (dl_no_of_rows+1),column=(i+1)).value= (msglist1[keylist[i]])
for i in range(len(msglist1), len(keylist)):
    dl_sheet_obj.cell(row = (dl_no_of_rows+1),column=(i+1)).value= (msglist2[keylist[i]])
    # dl_sheet_obj.cell(row = (dl_no_of_rows+1),column=(i+2)).value= (msglist1[keylist[i+2]])
# for i in range(len(keylist[1])):
#     dl_sheet_obj.cell(row = (dl_no_of_rows+1),column=(i+53)).value= (msglist2[keylist[1][i]])            

dl_sheet_obj.cell(row = (dl_no_of_rows+1),column=1).value= str(Serial_number)
dl_sheet_obj.cell(row = (dl_no_of_rows+1),column=2).value= str(bms_battery_code)
dl_obj.save(dl_path)





        