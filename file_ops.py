import os
import openpyxl
class fileoperations:
    def excel_read_file(filename):
        wb = openpyxl.load_workbook(filename)
        # Sheet = wb.sheetnames
        # print(wb.active.title)
        sh1 = wb['Sheet1']
        row = sh1.max_row
        column = sh1.max_column
        local_excel_data = []
        for i in range(1, row+1):
            row_data = []
            for j in range(1, column+1): 
                row_data.append(sh1.cell(row=i, column=j).value)
            local_excel_data.append(row_data)
        return local_excel_data
    def remove_file(filename):
        script_dir = os.getcwd()
        #  = input("Enter the name of file : ")
        print(filename)
        if os.path.exists(filename):
            try:
                os.remove(os.path.normcase(os.path.join(script_dir, filename)))
                return 
            except Exception as err:
                return 
        else:
            return "file is not exist"
            # return ("file is not exist")
    